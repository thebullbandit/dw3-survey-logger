"""
Boxel Sheet Exporter
====================

Exports boxel size survey data (commander name + highest-numbered system)
to an Excel file from the boxel_entries table.

Columns: Timestamp, Commander Name, Highest System in Boxel, Column 3
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Dict, Any, List

import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_boxel_sheet(
    entries: Iterable[Dict[str, Any]],
    output_dir: Path,
    cmdr_name: str = "UnknownCMDR",
) -> Optional[Path]:
    """
    Export boxel entries to an XLSX file.

    Args:
        entries: Iterable of dicts from observer_storage.get_boxel_entries()
        output_dir: Directory to write the file into
        cmdr_name: Commander name for the filename

    Returns:
        Path to created file, or None if no boxel data found
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    rows = [e for e in entries if e and (e.get("boxel_highest_system") or "").strip()]
    if not rows:
        return None

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Boxel Size"

    # --- Styles ---
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=10)
    data_font_ts = Font(name="Calibri", size=10, color="555555")
    even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="D9E2EC"),
    )

    # --- Title block ---
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "DW3 Stellar Properties Boxel Size"
    title_cell.font = title_font
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = f"CMDR {cmdr_name}  ·  Exported {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    sub_cell.font = subtitle_font
    ws.row_dimensions[2].height = 20

    # Blank spacer row
    ws.row_dimensions[3].height = 8

    # --- Header row (row 4) ---
    HEADER_ROW = 4
    headers = ["Timestamp", "Commander Name", "Highest System in Boxel", "Column 3"]
    col_widths = [24, 28, 42, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[HEADER_ROW].height = 24

    # --- Data rows ---
    DATA_START = HEADER_ROW + 1
    for i, d in enumerate(rows):
        row_num = DATA_START + i

        # Timestamp
        ts_str = d.get("created_at_utc") or ""
        try:
            if ts_str.endswith("Z"):
                ts_str = ts_str[:-1] + "+00:00"
            ts = datetime.fromisoformat(ts_str).replace(tzinfo=None)
            ts_cell = ws.cell(row=row_num, column=1, value=ts)
            ts_cell.number_format = "YYYY-MM-DD HH:MM:SS"
        except Exception:
            ts_cell = ws.cell(row=row_num, column=1, value=ts_str)
        ts_cell.font = data_font_ts

        # Commander name
        cmdr_cell = ws.cell(row=row_num, column=2, value=d.get("cmdr_name") or cmdr_name)
        cmdr_cell.font = data_font

        # Highest system
        sys_cell = ws.cell(row=row_num, column=3, value=(d.get("boxel_highest_system") or "").strip())
        sys_cell.font = data_font

        # Column 3 (empty)
        ws.cell(row=row_num, column=4).font = data_font

        # Alternating row shading + subtle bottom border
        if i % 2 == 0:
            for col_idx in range(1, 5):
                ws.cell(row=row_num, column=col_idx).fill = even_fill

        for col_idx in range(1, 5):
            ws.cell(row=row_num, column=col_idx).border = thin_border

    # Freeze header row so it stays visible when scrolling
    ws.freeze_panes = f"A{DATA_START}"

    # Auto-filter on the header row
    ws.auto_filter.ref = f"A{HEADER_ROW}:D{HEADER_ROW + len(rows)}"

    # --- Save ---
    safe_cmdr = re.sub(r"[^A-Za-z0-9_-]", "_", cmdr_name or "UnknownCMDR")
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"DW3_Stellar_Properties_Boxels_{safe_cmdr}_{ts}.xlsx"
    file_path = output_dir / filename

    wb.save(file_path)
    return file_path
