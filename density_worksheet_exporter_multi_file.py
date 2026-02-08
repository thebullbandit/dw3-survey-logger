"""
Density Worksheet Exporter - Multi-File Version
================================================

Creates separate DW3 "Stellar Density Scan Worksheet" Excel files for each sample.
Each sample (1000ly span with ~20 systems) gets its own .xlsx file.

Modified from original to support:
- One .xlsx file per sample_index
- Files named like "DW3_Stellar_Density_CMDR_Sample_01.xlsx"
- Preserves all original functionality per file
- **NEW**: Formula recalculation to ensure charts display correctly
"""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Dict, Any, List
from collections import defaultdict

import sys
import openpyxl
import re
import subprocess
import logging

from observer_models import SurveyType


def resource_path(*parts: str) -> Path:
    """
    Returns an absolute Path to a bundled resource.
    Works for: PyInstaller onefile + normal python runs.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base = Path(sys._MEIPASS)  # type: ignore[attr-defined]
    else:
        base = Path(__file__).resolve().parent
    return base.joinpath(*parts)


template_path = resource_path("templates", "Stellar Density Scan Worksheet.xlsx")


def _safe_parse_iso(ts: str) -> Optional[datetime]:
    try:
        # Handles both "...Z" and "+00:00"
        if ts.endswith("Z"):
            ts = ts[:-1] + "+00:00"
        return datetime.fromisoformat(ts)
    except Exception:
        return None


def _recalculate_formulas(file_path: Path, timeout: int = 30) -> bool:
    """
    Recalculates formulas in an Excel file using LibreOffice.
    
    Args:
        file_path: Path to the Excel file
        timeout: Maximum time in seconds to wait for recalculation
        
    Returns:
        True if recalculation succeeded, False otherwise
    """
    # Look for recalc.py script in common locations
    script_locations = [
        resource_path("scripts", "recalc.py"),
        Path("scripts/recalc.py"),
        Path(__file__).parent / "scripts" / "recalc.py",
    ]
    
    recalc_script = None
    for loc in script_locations:
        if loc.exists():
            recalc_script = loc
            break
    
    if not recalc_script:
        logging.warning(
            "Formula recalculation script not found. "
            "Charts may not display correctly until opened in Excel. "
            f"Searched locations: {[str(loc) for loc in script_locations]}"
        )
        return False
    
    try:
        result = subprocess.run(
            [sys.executable, str(recalc_script), str(file_path), str(timeout)],
            capture_output=True,
            text=True,
            timeout=timeout + 5
        )
        
        if result.returncode == 0:
            logging.info(f"Successfully recalculated formulas for {file_path.name}")
            return True
        else:
            logging.warning(
                f"Formula recalculation returned error code {result.returncode} "
                f"for {file_path.name}: {result.stderr}"
            )
            return False
            
    except subprocess.TimeoutExpired:
        logging.warning(f"Formula recalculation timed out for {file_path.name}")
        return False
    except Exception as e:
        logging.warning(f"Could not recalculate formulas for {file_path.name}: {e}")
        return False


def export_density_worksheet_from_notes_multi_file(
    notes: Iterable[Any],
    template_path: Path,
    output_path: Path,
    cmdr_name: str = "UnknownCMDR",
    sample_tag: str = "",
    z_bin: Optional[int] = None,
    sheet_name: str = "Blank CW",
    recalculate: bool = True,
    survey_type: Optional[SurveyType] = None,
) -> List[Path]:
    """
    Export density worksheet data to Excel files.

    Args:
        notes: Iterable of note objects containing density scan data
        template_path: Path to template Excel file (ignored if survey_type specified)
        output_path: Output directory or file path
        cmdr_name: Commander name for the worksheet header
        sample_tag: Optional tag for the sample
        z_bin: Optional Z-axis bin number
        sheet_name: Name of the sheet to use in the template
        recalculate: If True, attempt to recalculate formulas after save (default: True)
        survey_type: If specified, filter notes by this survey type and use matching template

    Returns:
        List of paths to created Excel files
    """
    # Select template based on survey type
    if survey_type == SurveyType.LOGARITHMIC_DENSITY:
        template_path = resource_path("templates", "Logarithmic Stellar Density Scan Worksheet.xlsx")
    else:
        template_path = resource_path("templates", "Stellar Density Scan Worksheet.xlsx")
    output_path = Path(output_path)
    _cmdr = cmdr_name

    # Normalize notes into dicts
    normalized: List[Dict[str, Any]] = []
    for n in notes:
        if n is None:
            continue
        if isinstance(n, dict):
            d = n
        else:
            try:
                d = asdict(n)
            except Exception:
                d = dict(n)
        normalized.append(d)

    # Filter to valid density rows
    all_rows: List[Dict[str, Any]] = []
    for d in normalized:
        if not (d.get("system_name") or ""):
            continue
        if d.get("max_distance") is None:
            continue
        rs = str(d.get("record_status", "")).lower()
        if rs in ("amended", "deleted", "inactive"):
            continue
        # Filter by survey_type if specified
        if survey_type is not None:
            note_survey_type = d.get("survey_type")
            # Handle both enum and string values
            if isinstance(note_survey_type, SurveyType):
                if note_survey_type != survey_type:
                    continue
            elif isinstance(note_survey_type, str):
                if note_survey_type != survey_type.value:
                    continue
            else:
                # If survey_type not set on note, include it for backward compatibility
                # (old notes without survey_type are assumed to be regular density)
                if survey_type != SurveyType.REGULAR_DENSITY:
                    continue
        all_rows.append(d)

    if not all_rows:
        survey_name = survey_type.value if survey_type else "any"
        raise ValueError(
            f"No completed samples to export for survey type '{survey_name}'. "
            "Save at least one observation with density data before exporting."
        )

    # Group rows by sample_index
    samples_dict: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        try:
            samples_dict[int(row.get("sample_index"))].append(row)
        except Exception:
            pass

    if not samples_dict:
        raise ValueError("No valid sample_index found in data. Cannot group samples.")

    created_files: List[Path] = []

    safe_cmdr = re.sub(r"[^A-Za-z0-9_-]", "_", (_cmdr or "UnknownCMDR"))
    z_part = f"_Z{int(z_bin)}" if z_bin is not None else ""

    base_dir = output_path.parent if output_path.suffix.lower() == ".xlsx" else output_path
    base_dir.mkdir(parents=True, exist_ok=True)

    for sample_idx in sorted(samples_dict.keys()):
        rows = samples_dict[sample_idx]

        # Sort rows by timestamp + system_index
        def sort_key(d: Dict[str, Any]):
            ts = _safe_parse_iso(str(d.get("timestamp_utc") or ""))
            try:
                si = int(d.get("system_index") or 0)
            except Exception:
                si = 0
            return (ts or datetime.min, si)

        rows.sort(key=sort_key)

        wb = openpyxl.load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        START_ROW = 6

        # 🔒 Static Z Sample column (B): 0..1000 step 50
        static_z = list(range(0, 1001, 50))
        for i, z in enumerate(static_z):
            ws.cell(START_ROW + i, 2).value = z

        # Header: CMDR name and date
        ws["A1"].value = f"CMDR {_cmdr or 'UnknownCMDR'} - DW3 Stellar Density Scans"
        if rows:
            dt = _safe_parse_iso(str(rows[0].get("timestamp_utc") or ""))
            if dt:
                ws["B2"].value = dt.date().isoformat()

        # Write data rows
        for i, d in enumerate(rows):
            r = START_ROW + i

            ws.cell(r, 1).value = d.get("system_name") or ""     # A System
            # Column B intentionally NOT written (static Z values)

            ws.cell(r, 3).value = d.get("system_count")         # C System Count

            corrected = d.get("corrected_n")
            if corrected is None:
                sc = d.get("system_count")
                corrected = (sc + 1) if sc is not None else None
            if corrected is not None:
                ws.cell(r, 4).value = corrected                 # D Corrected n

            ws.cell(r, 5).value = d.get("max_distance")         # E Max Distance

            # Calculate and set Rho value (column F)
            # This ensures LibreOffice has a cached value to display
            # The formula will recalculate when Excel/LibreOffice opens the file
            corrected_n = corrected if corrected is not None else None
            max_dist = d.get("max_distance")
            if corrected_n is not None and max_dist is not None:
                try:
                    import math
                    if corrected_n == 50:
                        # Rho = 50 / ((4*PI/3) * max_distance^3)
                        rho = 50 / ((4 * math.pi / 3) * (max_dist ** 3))
                    elif corrected_n < 50:
                        # Rho = corrected_n / ((4*pi/3) * 20^3)
                        rho = corrected_n / ((4 * math.pi / 3) * (20 ** 3))
                    else:
                        rho = None
                    
                    if rho is not None:
                        # Set the calculated value as cached value for the formula cell
                        cell_f = ws.cell(r, 6)  # Column F
                        # The cell already has the formula from the template
                        # We just need to make sure it has a cached value
                        # openpyxl doesn't support setting cached values directly,
                        # so we'll overwrite with the calculated value
                        ws.cell(r, 6).value = rho
                except Exception as e:
                    # If calculation fails, leave the formula as is
                    pass

            sp = d.get("star_pos") or (None, None, None)
            try:
                x, y, z = sp
            except Exception:
                x = y = z = None

            ws.cell(r, 7).value = x                              # G X
            ws.cell(r, 8).value = y                              # H Y
            ws.cell(r, 9).value = z                              # I Z

            # Calculate Distance from Sol (column J) and R from Core (column K)
            # These replace formula-based calculations to avoid #NAME? errors
            if x is not None and y is not None and z is not None:
                try:
                    import math
                    # Distance from Sol (Sol is at 0, 0, 0)
                    dist_from_sol = math.sqrt(x**2 + y**2 + z**2)
                    ws.cell(r, 10).value = dist_from_sol          # J Dist from Sol

                    # R from Galactic Core (Sagittarius A* is approximately at -25.8, -20.5, 25900)
                    # Using simplified galactic center coordinates
                    gc_x, gc_y, gc_z = 0, 0, 25900  # Simplified: just Z offset
                    r_from_core = math.sqrt(x**2 + y**2 + (z - gc_z)**2)
                    ws.cell(r, 11).value = r_from_core            # K R from Core
                except Exception:
                    pass  # Leave cells empty if calculation fails

        # Clear unused rows (rows after the data) to prevent #DIV/0! and #NAME? errors
        # The template has formulas for rows 6-26, but we may only use some of them
        last_data_row = START_ROW + len(rows) - 1
        for r in range(last_data_row + 1, START_ROW + 21):  # Clear rows up to row 26
            # Clear all columns except B (which has static Z values)
            # Columns: A=1, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11
            for col in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
                cell = ws.cell(r, col)
                cell.value = None
                # Also clear any formula
                if hasattr(cell, '_value'):
                    cell._value = None

        # Force Excel to recalculate formulas when the file is opened
        # This ensures charts display correctly even without LibreOffice
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True

        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        # Include survey type in filename for clarity
        if survey_type == SurveyType.LOGARITHMIC_DENSITY:
            survey_prefix = "DW3_Logarithmic_Density"
        else:
            survey_prefix = "DW3_Regular_Density"
        filename = f"{survey_prefix}_{safe_cmdr}{z_part}_Sample_{sample_idx:02d}_{ts}.xlsx"
        file_path = base_dir / filename

        wb.save(file_path)
        created_files.append(file_path)
        
        # Attempt to recalculate formulas immediately if requested
        if recalculate:
            _recalculate_formulas(file_path)

    return created_files
