"""
Density Worksheet Exporter
==========================

Creates the DW3 "Stellar Density Scan Worksheet" Excel file and auto-fills
it from ObserverNotes collected by the logger.

Design goals:
- Use the existing DW3 worksheet template (keeps formulas intact)
- Only fill the minimum required fields (privacy-minimizing)
- Deterministic mapping: one row per observed system (DW3 manual workflow)

Inputs populated per row:
A: System
B: Z Sample (Z-bin)
C: System Count
D: Corrected n
E: Max Distance
G: X
H: Z
I: Y

Everything else is left to the template formulas (Rho, integrals, totals).
"""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Dict, Any, List

import openpyxl
import re


def _safe_parse_iso(ts: str) -> Optional[datetime]:
    try:
        # Handles both "...Z" and "+00:00"
        if ts.endswith("Z"):
            ts = ts[:-1] + "+00:00"
        return datetime.fromisoformat(ts)
    except Exception:
        return None


def export_density_worksheet_from_notes(
    notes: Iterable[Any],
    template_path: Path,
    output_path: Path,
    cmdr_name: str = "UnknownCMDR",
    sample_tag: str = "",
    z_bin: Optional[int] = None,
    sheet_name: str = "Blank CW",
) -> Path:
   
    template_path = Path(template_path)
    output_path = Path(output_path)
    _cmdr = cmdr_name


    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    # Normalize notes into dicts
    normalized: List[Dict[str, Any]] = []
    for n in notes:
        if n is None:
            continue
        if isinstance(n, dict):
            d = n
        else:
            try:
                d = asdict(n)
            except Exception:
                d = dict(n)
        normalized.append(d)

    # Fill header date (B2) if we have timestamps
    if normalized:
        dt = _safe_parse_iso(str(normalized[0].get("timestamp_utc") or ""))
        if dt:
            ws["B2"].value = dt.date().isoformat()

    # Filter to per-system density entries (system name + max_distance present)
    rows: List[Dict[str, Any]] = []
    for d in normalized:
        if not (d.get("system_name") or ""):
            continue
        if d.get("max_distance") is None:
            continue
        # If a record_status is present in payload, skip non-active
        rs = str(d.get("record_status", "")).lower()
        if rs in ("amended", "deleted", "inactive"):
            continue
        rows.append(d)

    # Sort by z_bin, then timestamp, then system_index for stable ordering
    def sort_key(d: Dict[str, Any]):
        try:
            z = int(d.get("z_bin", 0) or 0)
        except Exception:
            z = 0
        ts = _safe_parse_iso(str(d.get("timestamp_utc") or ""))
        try:
            si = int(d.get("system_index") or 0)
        except Exception:
            si = 0
        return (z, ts or datetime.min, si)

    rows.sort(key=sort_key)

    # Build an index of existing template rows by Z Sample (col B).
    z_to_rows: Dict[int, List[int]] = {}
    for r in range(6, min(ws.max_row, 2000) + 1):
        b = ws.cell(r, 2).value
        try:
            zb = int(b) if b is not None and str(b).strip() != "" else None
        except Exception:
            zb = None
        if zb is not None:
            z_to_rows.setdefault(zb, []).append(r)

    def find_row_for_z(zb: int) -> int:
        # Prefer existing row with matching Z Sample and empty System cell (A)
        for r in z_to_rows.get(zb, []):
            a = ws.cell(r, 1).value
            if not (str(a).strip() if a is not None else ""):
                return r
        # Otherwise append at bottom (avoid insert_rows which may break formulas)
        return ws.max_row + 1

    for d in rows:
        try:
            zb = int(d.get("z_bin") or 0)
        except Exception:
            zb = 0
        r = find_row_for_z(zb)

        ws.cell(r, 1).value = d.get("system_name") or ""      # A System
        ws.cell(r, 2).value = zb                               # B Z Sample

        # C System Count: CMDR-entered from NAV (required by DW3). Copy exactly as stored.
        ws.cell(r, 3).value = d.get("system_count")

        # D Corrected n: only if CMDR entered it in overlay
        if d.get("corrected_n") is not None:
            ws.cell(r, 4).value = d.get("corrected_n")          # D

        ws.cell(r, 5).value = d.get("max_distance")             # E Max Distance

        # StarPos mapping: logger stores (x, y, z) but worksheet expects X, Z, Y
        sp = d.get("star_pos") or (None, None, None)
        try:
            x, y, z = sp
        except Exception:
            x = y = z = None

        ws.cell(r, 7).value = x                                 # G X
        ws.cell(r, 8).value = z                                 # H Z
        ws.cell(r, 9).value = y                                 # I Y

    # --------------------------------------------------------------------
    # Output filename normalization (DW3-friendly)
    # - Callers sometimes pass a full file path like:
    #     DW3_Stellar_Density_Worksheet_YYYYMMDD_HHMMSS.xlsx
    # - DW3 wants CMDR in the filename. We rewrite the name while keeping the
    #   same directory and timestamp.
    # --------------------------------------------------------------------
        safe_cmdr = re.sub(r"[^A-Za-z0-9_-]", "_", (_cmdr or "UnknownCMDR"))
    safe_sample = re.sub(r"[^A-Za-z0-9_-]", "_", (sample_tag or "")).strip("_")
    z_part = f"_Z{int(z_bin)}" if z_bin is not None else ""
    sample_part = f"_{safe_sample}" if safe_sample else ""

    final_path = Path(output_path)

    # If output_path is a directory or has no .xlsx suffix, generate a filename.
    if final_path.suffix.lower() != ".xlsx":
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        final_path = final_path / f"DW3_Stellar_Density_{safe_cmdr}{z_part}{sample_part}_{ts}.xlsx"
    else:
        # Try to preserve any timestamp already in the filename.
        m = re.search(r"(\d{8}_\d{6})", final_path.name)
        ts = m.group(1) if m else datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        final_path = final_path.with_name(f"DW3_Stellar_Density_{safe_cmdr}{z_part}{sample_part}_{ts}.xlsx")

    final_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(final_path)
    return final_path
